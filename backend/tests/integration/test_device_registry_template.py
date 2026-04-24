"""Tests for device-registry bulk template XLSX generation."""

import io

import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy import delete


pytestmark = pytest.mark.asyncio


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def test_bulk_template_returns_xlsx_with_two_sheets(
    async_client: AsyncClient, admin_token: str
):
    """RED: verify XLSX Content-Type, filename, and two-sheet structure."""
    resp = await async_client.get(
        "/device-registry/bulk/template",
        headers=_auth(admin_token),
    )

    assert resp.status_code == 200
    assert (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        in resp.headers.get("content-type", "")
    ), f"Expected XLSX content-type, got: {resp.headers.get('content-type')}"
    assert (
        "device_registry_bulk_template.xlsx"
        in resp.headers.get("content-disposition", "")
    ), f"Expected .xlsx filename, got: {resp.headers.get('content-disposition')}"

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    sheet_names = wb.sheetnames
    assert len(sheet_names) == 2, f"Expected 2 sheets, got {sheet_names}"
    assert "Template" in sheet_names
    assert "Categories" in sheet_names


async def test_template_sheet_has_correct_headers_and_example_rows(
    async_client: AsyncClient, admin_token: str
):
    """RED: Template sheet must have required headers and at least one example row."""
    resp = await async_client.get(
        "/device-registry/bulk/template",
        headers=_auth(admin_token),
    )
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Template"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == [
        "serial_number",
        "name",
        "description",
        "category_id",
        "ip_address",
        "mac_address",
    ], f"Template headers mismatch: {headers}"

    # At least one example row (row 2 must have non-empty serial_number)
    row2 = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    assert row2[0] is not None and str(row2[0]).strip() != "", (
        "Template sheet row 2 must have an example serial_number"
    )


async def test_categories_sheet_populated_from_nas_category_table(
    async_client: AsyncClient,
    admin_token: str,
    test_db,
):
    """Categories sheet must contain rows from NasCategory table."""
    from app.models.models import NasCategory

    # Seed two categories
    cat1 = NasCategory(name="Residential AP", description="Standard home routers")
    cat2 = NasCategory(name="Enterprise AP", description="High-capacity sector")
    test_db.add_all([cat1, cat2])
    await test_db.commit()

    try:
        resp = await async_client.get(
            "/device-registry/bulk/template",
            headers=_auth(admin_token),
        )
        assert resp.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Categories"]

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == ["id", "name", "description"], f"Categories headers mismatch: {headers}"

        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        row_names = sorted(r[1] for r in data_rows if r[1] is not None)
        assert row_names == ["Enterprise AP", "Residential AP"], (
            f"Categories sheet must have both seeded categories, got: {row_names}"
        )
    finally:
        # Clean up so subsequent tests don't see these categories
        await test_db.execute(
            delete(NasCategory).where(NasCategory.name.in_(["Residential AP", "Enterprise AP"]))
        )
        await test_db.commit()


async def test_categories_sheet_empty_when_no_categories_exist(
    async_client: AsyncClient, admin_token: str
):
    """Categories sheet should be created with headers only when DB has no NasCategory rows."""
    resp = await async_client.get(
        "/device-registry/bulk/template",
        headers=_auth(admin_token),
    )
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Categories"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["id", "name", "description"]

    # Verify the sheet has exactly 1 row (the header) when DB is empty
    all_rows = list(ws.iter_rows(values_only=True))
    assert len(all_rows) == 1, (
        f"Categories sheet should have exactly 1 row (header only) when DB is empty, got {len(all_rows)} rows"
    )
